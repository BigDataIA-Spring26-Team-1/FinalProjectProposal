from __future__ import annotations

import asyncio
import base64
import logging
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import UTC, datetime, timedelta
from io import BytesIO
from typing import Any

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .config import Settings
from .models import (
    DashboardPayload,
    DatasetProfile,
    DatasetSourceProfile,
    FeedEntry,
    SourceCard,
    SourceSnapshot,
    StackStatusItem,
)

PROPOSAL_HIGHLIGHTS = [
    "The source universe spans 45.5+ GB across 7 manufacturing-relevant source domains.",
    "React operations console supports plant layout editing, simulation, and scenario comparison.",
    "Poetry-managed FastAPI services consolidate external data into a unified dashboard payload.",
    "Live source adapters cover compliance, market, engineering, environmental, and equipment signals.",
    "Environment-based credential management keeps provider keys out of the frontend.",
]

DATASET_SOURCE_PROFILES = [
    DatasetSourceProfile(
        name="SEC EDGAR 10-K/10-Q Filings",
        acquisition="sec-edgar-downloader + PDF text extraction + HTTP fetch",
        estimated_size="~8 GB",
        challenge=(
            "Structured financial filings require parsing annual and quarterly reports, "
            "extracting capex, equipment depreciation, maintenance language, and production clues "
            "from long-form documents."
        ),
    ),
    DatasetSourceProfile(
        name="OSHA Inspection and Accident Database",
        acquisition="REST API pagination or official OSHA.gov scraping fallback",
        estimated_size="~12 GB",
        challenge=(
            "Millions of workplace inspection and violation rows need incremental ingestion, "
            "normalization, and cross-referencing for facility-level risk analysis."
        ),
    ),
    DatasetSourceProfile(
        name="DOE Industrial Assessment Center Audits",
        acquisition="Bulk government download with schema joins",
        estimated_size="~2 GB",
        challenge=(
            "Audit workbooks include equipment types, energy consumption, recommended improvements, "
            "and measured savings that calibrate factory efficiency scenarios."
        ),
    ),
    DatasetSourceProfile(
        name="Equipment Manufacturer Catalogs",
        acquisition="Async web scraping and document parsing",
        estimated_size="~15 GB",
        challenge=(
            "Industrial catalogs and manuals must be scraped, cleaned, and normalized across vendors "
            "to recover specs, power draw, maintenance intervals, parts, and technical guidance."
        ),
    ),
    DatasetSourceProfile(
        name="FRED + EIA Economic and Energy APIs",
        acquisition="REST APIs with refresh windows and incremental pulls",
        estimated_size="~500 MB",
        challenge=(
            "Economic indicators and electricity pricing feed the simulation with external demand, "
            "cost, and energy context rather than hardcoded assumptions."
        ),
    ),
    DatasetSourceProfile(
        name="EPA Toxic Release Inventory",
        acquisition="Bulk CSV and API enrichment",
        estimated_size="~5 GB",
        challenge=(
            "Environmental compliance data can be joined with operational records to build plant-level "
            "profiles for emissions, waste, and compliance cost modeling."
        ),
    ),
    DatasetSourceProfile(
        name="Engineering Forums Knowledge Base",
        acquisition="Large-scale forum scraping with classification",
        estimated_size="~3 GB",
        challenge=(
            "Forum content requires filtering, chunking, and indexing to isolate maintenance and "
            "failure-pattern knowledge that complements manuals and equipment sources."
        ),
    ),
]

DATASET_PROFILE = DatasetProfile(
    total_estimated_size="45.5+ GB",
    source_count=7,
    join_keys=["facility ID", "NAICS code", "equipment type"],
    engineering_note=(
        "This is a genuine big-data integration problem because each source uses a different acquisition strategy, "
        "schema, and refresh pattern. Spark is the scale-out processing path for full ingestion, and Airflow is the "
        "orchestration path for scheduled pipelines and dependency management."
    ),
    sources=DATASET_SOURCE_PROFILES,
)

logger = logging.getLogger(__name__)


def _now_iso() -> str:
    return datetime.now(UTC).isoformat()


def _extract_records(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]

    if isinstance(payload, dict):
        for key in ("data", "results", "items", "records"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]

        for value in payload.values():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                return [item for item in value if isinstance(item, dict)]

    return []


def _pick_fields(record: dict[str, Any], limit: int = 6) -> dict[str, Any]:
    picked: dict[str, Any] = {}
    for key, value in record.items():
        if value in (None, "", [], {}):
            continue
        picked[key] = value
        if len(picked) >= limit:
            break
    return picked


def _coerce_float(value: Any) -> float | None:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return numeric if numeric > -900 else None


def _coerce_int(value: Any) -> int | None:
    cleaned = re.sub(r"[^\d\-]", "", str(value or ""))
    if not cleaned:
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").split())


def _describe_exception(exc: Exception) -> str:
    message = _clean_text(str(exc))
    if message:
        return message
    return exc.__class__.__name__


def _tone_for_state(state: str) -> str:
    if state == "online":
        return "ok"
    if state == "partial":
        return "warn"
    return "alert"


def _is_credential_warning(message: str | None) -> bool:
    lowered = (message or "").lower()
    return "key" in lowered and any(
        token in lowered
        for token in ("missing", "set ", "add ", "waiting for", "enable")
    )


def _friendly_warning_text(key: str, message: str | None) -> str | None:
    if not message:
        return None

    lowered = message.lower()
    if any(
        token in lowered
        for token in (
            "source is temporarily unreachable",
            "source request timed out",
            "source authentication failed",
            "source rate limit reached",
            "a secure connection to the source could not be established",
            "one or more vendor catalog pages blocked parsing",
            "feed is waiting for a department of labor api key",
            "feed is waiting for a federal reserve api key",
            "feed is waiting for an energy information administration api key",
            "serpapi search feed is waiting for an api key",
            "glassdoor feed is waiting for a rapidapi key",
            "this live feed is waiting for api credentials",
        )
    ):
        return message

    credential_messages = {
        "osha": "OSHA feed is waiting for a Department of Labor API key.",
        "fred": "FRED feed is waiting for a Federal Reserve API key.",
        "eia": "EIA feed is waiting for an Energy Information Administration API key.",
        "serpapi": "SerpAPI search feed is waiting for an API key.",
        "glassdoor": "Glassdoor feed is waiting for a RapidAPI key.",
    }

    if _is_credential_warning(message):
        return credential_messages.get(key, "This live feed is waiting for API credentials.")

    if any(
        token in lowered
        for token in (
            "getaddrinfo failed",
            "name or service not known",
            "temporary failure in name resolution",
            "nodename nor servname provided",
        )
    ):
        return "Source is temporarily unreachable from the current network."

    if "timed out" in lowered or "timeout" in lowered:
        return "Source request timed out before the provider responded."

    if any(token in lowered for token in ("401", "403", "unauthorized", "forbidden")):
        return "Source authentication failed. Verify the configured credentials."

    if any(token in lowered for token in ("429", "rate limit", "too many requests")):
        return "Source rate limit reached. Retry after the provider cooldown."

    if any(token in lowered for token in ("ssl", "tls", "certificate")):
        return "A secure connection to the source could not be established."

    if key == "catalogs":
        return "One or more vendor catalog pages blocked parsing or returned incomplete product data."

    return "Source is temporarily unavailable. Try refreshing later."


def _tone_for_snapshot(snapshot: SourceSnapshot) -> str:
    if snapshot.state == "offline" and _is_credential_warning(snapshot.warning):
        return "warn"
    return _tone_for_state(snapshot.state)


def _display_text(snapshot: SourceSnapshot) -> str:
    friendly_warning = _friendly_warning_text(snapshot.key, snapshot.warning)
    if snapshot.state == "offline" and friendly_warning:
        return friendly_warning
    if snapshot.state == "partial" and friendly_warning:
        return f"{snapshot.summary} {friendly_warning}"
    return snapshot.summary


def _source_volume(snapshot: SourceSnapshot) -> str:
    if snapshot.state == "offline":
        if _is_credential_warning(snapshot.warning):
            return "Waiting for credentials"
        return "Sync paused"
    if snapshot.count:
        return f"{snapshot.count} sample records"
    return "No fresh records"


def _status_name(source: SourceSnapshot) -> str:
    replacements = {
        "sec": "EDGAR",
        "osha": "OSHA",
        "itac": "ITAC",
        "epa": "EPA TRI",
        "nasa": "NASA",
        "fred": "FRED",
        "eia": "EIA",
        "serpapi": "SerpAPI",
        "glassdoor": "Glassdoor",
        "stackexchange": "StackEx",
        "reddit": "Reddit",
        "catalogs": "Catalogs",
    }
    return replacements.get(source.key, source.name)


def _parse_itac_zip(content: bytes, sample_rows: int) -> list[dict[str, Any]]:
    with zipfile.ZipFile(BytesIO(content)) as archive:
        workbook_name = next(
            (name for name in archive.namelist() if name.lower().endswith(".xlsx")),
            None,
        )
        if workbook_name is None:
            return []

        workbook_bytes = archive.read(workbook_name)

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = next(
        (
            workbook[name]
            for name in workbook.sheetnames
            if "recommend" in name.lower() or "assessment" in name.lower()
        ),
        workbook[workbook.sheetnames[0]],
    )

    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, [])
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    indexes = [index for index, name in enumerate(headers) if name][:8]

    samples: list[dict[str, Any]] = []
    for row in rows:
        if not row or not any(row):
            continue
        record = {
            headers[index] or f"column_{index + 1}": row[index]
            for index in indexes
            if index < len(row) and row[index] not in (None, "")
        }
        if record:
            samples.append(record)
        if len(samples) >= sample_rows:
            break

    workbook.close()
    return samples


def _find_osha_table(soup: BeautifulSoup) -> Any | None:
    for table in soup.find_all("table"):
        headers = {
            _clean_text(cell.get_text(" ", strip=True)).lower()
            for cell in table.find_all("th")
        }
        if {"standard", "citations", "inspections"}.issubset(headers):
            return table
    return None


def _parse_osha_cited_standards(
    html: str,
    naics_code: str,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None, str | None]:
    soup = BeautifulSoup(html, "html.parser")
    table = _find_osha_table(soup)
    if table is None:
        return [], None, None

    page_text = _clean_text(soup.get_text(" ", strip=True))
    period_match = re.search(
        r"issued during the period\s+([A-Za-z]+\s+\d{4}\s+through\s+[A-Za-z]+\s+\d{4})",
        page_text,
        re.IGNORECASE,
    )
    period = period_match.group(1) if period_match else None

    records: list[dict[str, Any]] = []
    totals: dict[str, Any] | None = None
    for row in table.find_all("tr"):
        cells = row.find_all(["th", "td"])
        values = [_clean_text(cell.get_text(" ", strip=True)) for cell in cells]
        if len(values) < 5:
            continue

        first_cell = values[0].lower()
        if first_cell == "standard":
            continue

        standard_code, citations, inspections, penalty, description = values[:5]
        row_payload = {
            "naics_code": naics_code,
            "standard_code": standard_code,
            "citations": _coerce_int(citations),
            "inspections": _coerce_int(inspections),
            "penalty_usd": _coerce_int(penalty),
            "description": description,
        }

        if first_cell == "total":
            totals = row_payload
            continue

        records.append(row_payload)

    return records, totals, period


class DashboardAggregator:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self._cache: DashboardPayload | None = None
        self._cache_expires_at: datetime | None = None

    def get_cached_dashboard(self) -> DashboardPayload | None:
        return self._cache

    def build_placeholder_dashboard(self) -> DashboardPayload:
        return DashboardPayload(
            fetched_at=_now_iso(),
            stackStatus=[],
            liveFeed=[],
            proposalHighlights=PROPOSAL_HIGHLIGHTS,
            datasetProfile=DATASET_PROFILE,
            sourceCards=[],
            sourceSnapshots=[],
        )

    async def build_dashboard(self, force_refresh: bool = False) -> DashboardPayload:
        if (
            not force_refresh
            and self._cache is not None
            and self._cache_expires_at is not None
            and datetime.now(UTC) < self._cache_expires_at
        ):
            return self._cache

        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            snapshots = await asyncio.gather(
                self._guard("sec", "SEC EDGAR", self.fetch_sec(client)),
                self._guard("osha", "OSHA", self.fetch_osha(client)),
                self._guard("itac", "ITAC / DOE", self.fetch_itac(client)),
                self._guard("epa", "EPA TRI", self.fetch_epa_tri(client)),
                self._guard("nasa", "NASA POWER", self.fetch_nasa_power(client)),
                self._guard("fred", "FRED", self.fetch_fred(client)),
                self._guard("eia", "EIA", self.fetch_eia(client)),
                self._guard("serpapi", "SerpAPI Search", self.fetch_serpapi(client)),
                self._guard("glassdoor", "Glassdoor", self.fetch_glassdoor(client)),
                self._guard("stackexchange", "Stack Exchange", self.fetch_stackexchange(client)),
                self._guard("reddit", "Reddit", self.fetch_reddit(client)),
                self._guard("catalogs", "Equipment Catalogs", self.fetch_catalogs(client)),
            )

        dashboard = self._assemble_dashboard(snapshots)
        self._cache = dashboard
        self._cache_expires_at = datetime.now(UTC) + timedelta(
            seconds=self.settings.dashboard_cache_ttl_seconds,
        )
        return dashboard

    async def _guard(
        self,
        key: str,
        name: str,
        coro: Any,
    ) -> SourceSnapshot:
        try:
            return await coro
        except Exception as exc:
            logger.warning("%s fetch failed: %s", name, exc)
            return SourceSnapshot(
                key=key,
                name=name,
                state="offline",
                summary=f"{name} sync is paused.",
                warning=_friendly_warning_text(key, str(exc)),
                updated_at=_now_iso(),
            )

    async def fetch_sec(
        self,
        client: httpx.AsyncClient,
        tickers: list[str] | None = None,
        snapshot_key: str = "sec",
        snapshot_name: str = "SEC EDGAR",
    ) -> SourceSnapshot:
        headers = {"User-Agent": self.settings.sec_user_agent}
        ticker_map_response = await client.get("https://www.sec.gov/files/company_tickers.json", headers=headers)
        ticker_map_response.raise_for_status()
        ticker_payload = ticker_map_response.json()
        ticker_map = {
            entry["ticker"].upper(): str(entry["cik_str"]).zfill(10)
            for entry in ticker_payload.values()
            if isinstance(entry, dict) and entry.get("ticker") and entry.get("cik_str")
        }

        selected_tickers = [ticker.upper() for ticker in (tickers or self.settings.sec_ticker_list)][: self.settings.chat_market_research_limit]

        records: list[dict[str, Any]] = []
        for ticker in selected_tickers:
            cik = ticker_map.get(ticker)
            if not cik:
                continue
            submissions_url = f"https://data.sec.gov/submissions/CIK{cik}.json"
            response = await client.get(submissions_url, headers=headers)
            response.raise_for_status()
            payload = response.json()
            recent = payload.get("filings", {}).get("recent", {})
            forms = recent.get("form", [])
            accession_numbers = recent.get("accessionNumber", [])
            filing_dates = recent.get("filingDate", [])
            primary_documents = recent.get("primaryDocument", [])
            if not forms:
                continue
            records.append(
                {
                    "company": payload.get("name"),
                    "ticker": ticker,
                    "latest_form": forms[0],
                    "filing_date": filing_dates[0] if filing_dates else None,
                    "accession_number": accession_numbers[0] if accession_numbers else None,
                    "primary_document": primary_documents[0] if primary_documents else None,
                }
            )

        return SourceSnapshot(
            key=snapshot_key,
            name=snapshot_name,
            state="online" if records else "partial",
            summary=f"Fetched {len(records)} live filing snapshots for {', '.join(selected_tickers)}.",
            count=len(records),
            records=records,
            updated_at=_now_iso(),
        )

    async def fetch_osha(self, client: httpx.AsyncClient) -> SourceSnapshot:
        if self.settings.dol_api_key:
            headers = {"X-API-KEY": self.settings.dol_api_key}
            response = await client.get(
                f"{self.settings.dol_inspection_url}/limit/{self.settings.osha_limit}",
                headers=headers,
            )
            response.raise_for_status()
            records = [
                _pick_fields(record)
                for record in _extract_records(response.json())[: self.settings.osha_limit]
            ]

            return SourceSnapshot(
                key="osha",
                name="OSHA",
                state="online" if records else "partial",
                summary=f"Pulled {len(records)} recent OSHA inspection rows from the DOL API.",
                count=len(records),
                records=records,
                updated_at=_now_iso(),
            )

        all_records: list[dict[str, Any]] = []
        aggregate_citations = 0
        aggregate_inspections = 0
        observed_periods: list[str] = []

        for naics_code in self.settings.osha_naics_list:
            response = await client.get(
                self.settings.osha_public_citedstandards_url,
                params={
                    "p_esize": "",
                    "p_naics": naics_code,
                    "p_state": self.settings.osha_public_state_scope,
                },
            )
            response.raise_for_status()
            records, totals, period = _parse_osha_cited_standards(response.text, naics_code)
            all_records.extend(records)

            if totals:
                aggregate_citations += totals.get("citations") or 0
                aggregate_inspections += totals.get("inspections") or 0
            if period:
                observed_periods.append(period)

        ranked_records = sorted(
            all_records,
            key=lambda record: (
                record.get("citations") or 0,
                record.get("inspections") or 0,
            ),
            reverse=True,
        )[: self.settings.osha_limit]

        period_note = ""
        unique_periods = sorted(set(observed_periods))
        if unique_periods:
            period_note = f" Period {unique_periods[0]}."

        summary = (
            "Scraped OSHA manufacturing cited-standard data directly from OSHA.gov "
            f"for NAICS {', '.join(self.settings.osha_naics_list)}. "
            f"Top {len(ranked_records)} standards reflect about {aggregate_citations:,} citations "
            f"across {aggregate_inspections:,} inspections.{period_note}"
        )
        if not ranked_records:
            summary = "OSHA public cited-standard pages returned no manufacturing records."

        return SourceSnapshot(
            key="osha",
            name="OSHA",
            state="online" if ranked_records else "partial",
            summary=summary,
            count=len(ranked_records),
            records=ranked_records,
            updated_at=_now_iso(),
        )

    async def fetch_itac(self, client: httpx.AsyncClient) -> SourceSnapshot:
        page_response = await client.get(self.settings.itac_download_page_url)
        page_response.raise_for_status()
        page_text = page_response.text

        generated_match = re.search(r"Generated\s+([A-Za-z]{3}\s+\d{1,2},\s+\d{4})", page_text)
        assessments_match = re.search(r"ITAC Assessments\s+([\d,]+)", page_text)
        recommendations_match = re.search(r"Recommendations\s+([\d,]+)", page_text)

        zip_response = await client.get(self.settings.itac_zip_url)
        zip_response.raise_for_status()
        rows = await asyncio.to_thread(_parse_itac_zip, zip_response.content, self.settings.itac_sample_rows)

        summary_parts = []
        if generated_match:
            summary_parts.append(f"Generated {generated_match.group(1)}")
        if assessments_match:
            summary_parts.append(f"{assessments_match.group(1)} assessments")
        if recommendations_match:
            summary_parts.append(f"{recommendations_match.group(1)} recommendations")

        summary = ", ".join(summary_parts) if summary_parts else "Downloaded the current ITAC workbook."

        return SourceSnapshot(
            key="itac",
            name="ITAC / DOE",
            state="online" if rows else "partial",
            summary=summary,
            count=len(rows),
            records=[_pick_fields(row) for row in rows],
            updated_at=_now_iso(),
        )

    async def fetch_epa_tri(self, client: httpx.AsyncClient) -> SourceSnapshot:
        url = self.settings.epa_tri_url_template.format(
            state=self.settings.epa_state_abbr.upper(),
            end=max(self.settings.epa_tri_limit - 1, 0),
        )
        response = await client.get(url)
        response.raise_for_status()
        root = ET.fromstring(response.text)
        facilities = root.findall(".//tri_facility")
        records = []
        for facility in facilities[: self.settings.epa_tri_limit]:
            record = {
                child.tag.lower(): child.text
                for child in list(facility)[:8]
                if child.text not in (None, "")
            }
            if record:
                records.append(record)

        return SourceSnapshot(
            key="epa",
            name="EPA TRI",
            state="online" if records else "partial",
            summary=f"Fetched {len(records)} TRI facility records for {self.settings.epa_state_abbr.upper()}.",
            count=len(records),
            records=records,
            updated_at=_now_iso(),
        )

    async def fetch_nasa_power(self, client: httpx.AsyncClient) -> SourceSnapshot:
        end_date = datetime.now(UTC).date()
        start_date = end_date - timedelta(days=max(self.settings.nasa_lookback_days - 1, 0))
        response = await client.get(
            self.settings.nasa_power_daily_url,
            params={
                "parameters": ",".join(self.settings.nasa_power_parameter_list),
                "community": self.settings.nasa_power_community,
                "longitude": self.settings.nasa_longitude,
                "latitude": self.settings.nasa_latitude,
                "start": start_date.strftime("%Y%m%d"),
                "end": end_date.strftime("%Y%m%d"),
                "format": "JSON",
                "time-standard": "UTC",
            },
        )
        response.raise_for_status()
        payload = response.json()
        parameter_payload = payload.get("properties", {}).get("parameter", {})
        if not parameter_payload:
            return SourceSnapshot(
                key="nasa",
                name="NASA POWER",
                state="partial",
                summary="NASA POWER returned no parameter data.",
                updated_at=_now_iso(),
            )

        dates = sorted(
            {
                date_key
                for series in parameter_payload.values()
                if isinstance(series, dict)
                for date_key in series.keys()
            },
            reverse=True,
        )
        records: list[dict[str, Any]] = []
        for date_key in dates[: self.settings.nasa_lookback_days]:
            record = {
                "date": date_key,
                "solar_kwh_m2_day": _coerce_float(parameter_payload.get("ALLSKY_SFC_SW_DWN", {}).get(date_key)),
                "temp_max_c": _coerce_float(parameter_payload.get("T2M_MAX", {}).get(date_key)),
                "temp_min_c": _coerce_float(parameter_payload.get("T2M_MIN", {}).get(date_key)),
                "humidity_pct": _coerce_float(parameter_payload.get("RH2M", {}).get(date_key)),
                "wind_m_s": _coerce_float(parameter_payload.get("WS10M", {}).get(date_key)),
            }
            records.append(_pick_fields(record))

        latest = next(
            (
                record
                for record in records
                if record.get("temp_max_c") is not None or record.get("solar_kwh_m2_day") is not None
            ),
            records[0] if records else {},
        )
        latest_temp = latest.get("temp_max_c")
        latest_solar = latest.get("solar_kwh_m2_day")
        latest_details = []
        if latest_temp is not None:
            latest_details.append(f"max temp {latest_temp} C")
        if latest_solar is not None:
            latest_details.append(f"solar {latest_solar} kWh/m2/day")
        summary = (
            f"Fetched {len(records)} NASA POWER daily site-weather records for {self.settings.nasa_site_name}. "
            + (
                "Latest " + ", ".join(latest_details) + "."
                if latest_details
                else "Latest values are still pending in the near-real-time feed."
            )
            if records
            else f"NASA POWER returned no recent site-weather rows for {self.settings.nasa_site_name}."
        )
        if self.settings.nasa_api_key:
            summary += " NASA developer key is configured for future NASA Open API expansion."

        return SourceSnapshot(
            key="nasa",
            name="NASA POWER",
            state="online" if records else "partial",
            summary=summary,
            count=len(records),
            records=records,
            updated_at=_now_iso(),
        )

    async def fetch_fred(self, client: httpx.AsyncClient) -> SourceSnapshot:
        if not self.settings.fred_api_key:
            return SourceSnapshot(
                key="fred",
                name="FRED",
                state="offline",
                summary="FRED feed is waiting for credentials.",
                warning="Add a Federal Reserve API key to enable economic series retrieval.",
                updated_at=_now_iso(),
            )

        records: list[dict[str, Any]] = []
        warnings: list[str] = []
        for series_id in self.settings.fred_series_list:
            try:
                response = await client.get(
                    "https://api.stlouisfed.org/fred/series/observations",
                    params={
                        "api_key": self.settings.fred_api_key,
                        "series_id": series_id,
                        "file_type": "json",
                        "sort_order": "desc",
                        "limit": self.settings.fred_observation_limit,
                    },
                )
                response.raise_for_status()
            except httpx.HTTPStatusError as exc:
                warnings.append(f"{series_id} returned HTTP {exc.response.status_code}.")
                continue

            observations = response.json().get("observations", [])
            latest = observations[0] if observations else {}
            if latest:
                records.append(
                    {
                        "series_id": series_id,
                        "date": latest.get("date"),
                        "value": latest.get("value"),
                    }
                )

        return SourceSnapshot(
            key="fred",
            name="FRED",
            state="online" if records else "offline",
            summary=(
                f"Fetched {len(records)} FRED economic indicators."
                if records
                else "FRED sync is paused."
            ),
            count=len(records),
            records=records,
            warning=" ".join(warnings) if warnings else None,
            updated_at=_now_iso(),
        )

    async def fetch_eia(self, client: httpx.AsyncClient) -> SourceSnapshot:
        if not self.settings.eia_api_key:
            return SourceSnapshot(
                key="eia",
                name="EIA",
                state="offline",
                summary="EIA feed is waiting for credentials.",
                warning="Add an Energy Information Administration API key to enable energy price retrieval.",
                updated_at=_now_iso(),
            )

        response = await client.get(
            self.settings.eia_retail_sales_url,
            params={
                "api_key": self.settings.eia_api_key,
                "frequency": "monthly",
                "data[0]": "price",
                "facets[stateid][]": self.settings.eia_state_id,
                "facets[sectorid][]": "IND",
                "sort[0][column]": "period",
                "sort[0][direction]": "desc",
                "offset": 0,
                "length": self.settings.eia_length,
            },
        )
        response.raise_for_status()
        payload = response.json().get("response", {})
        data = payload.get("data", [])
        records = [_pick_fields(record) for record in data[: self.settings.eia_length]]

        return SourceSnapshot(
            key="eia",
            name="EIA",
            state="online" if records else "partial",
            summary=f"Fetched {len(records)} industrial electricity price records from EIA.",
            count=len(records),
            records=records,
            updated_at=_now_iso(),
        )

    async def fetch_serpapi(
        self,
        client: httpx.AsyncClient,
        query: str | None = None,
        snapshot_key: str = "serpapi",
        snapshot_name: str = "SerpAPI Search",
    ) -> SourceSnapshot:
        if not self.settings.serpapi_key:
            return SourceSnapshot(
                key=snapshot_key,
                name=snapshot_name,
                state="offline",
                summary="SerpAPI search feed is waiting for credentials.",
                warning="Add a SerpAPI key to enable live company and market search results.",
                updated_at=_now_iso(),
            )

        active_query = query or self.settings.serpapi_query
        response = await client.get(
            self.settings.serpapi_url,
            params={
                "engine": self.settings.serpapi_engine,
                "q": active_query,
                "api_key": self.settings.serpapi_key,
                "num": self.settings.serpapi_limit,
            },
        )
        response.raise_for_status()
        payload = response.json()
        items = payload.get("news_results") or payload.get("organic_results") or []
        records = [
            _pick_fields(
                {
                    "title": item.get("title"),
                    "source": item.get("source") or item.get("publication") or item.get("displayed_link"),
                    "snippet": item.get("snippet"),
                    "date": item.get("date"),
                    "link": item.get("link"),
                },
            )
            for item in items[: self.settings.serpapi_limit]
            if isinstance(item, dict)
        ]

        return SourceSnapshot(
            key=snapshot_key,
            name=snapshot_name,
            state="online" if records else "partial",
            summary=f"Fetched {len(records)} live SerpAPI results for '{active_query}'.",
            count=len(records),
            records=records,
            updated_at=_now_iso(),
        )

    async def fetch_glassdoor(
        self,
        client: httpx.AsyncClient,
        query: str | None = None,
        snapshot_key: str = "glassdoor",
        snapshot_name: str = "Glassdoor",
    ) -> SourceSnapshot:
        if not self.settings.rapidapi_key:
            return SourceSnapshot(
                key=snapshot_key,
                name=snapshot_name,
                state="offline",
                summary="Glassdoor feed is waiting for credentials.",
                warning="Add a RapidAPI key to enable live Glassdoor company results.",
                updated_at=_now_iso(),
            )

        active_query = query or self.settings.glassdoor_query
        try:
            response = await client.get(
                f"https://{self.settings.glassdoor_rapidapi_host}{self.settings.glassdoor_company_search_path}",
                headers={
                    "x-rapidapi-key": self.settings.rapidapi_key,
                    "x-rapidapi-host": self.settings.glassdoor_rapidapi_host,
                },
                params={
                    "query": active_query,
                    "limit": self.settings.glassdoor_limit,
                },
            )
            response.raise_for_status()
        except httpx.HTTPStatusError as exc:
            status_code = exc.response.status_code
            if status_code == 429:
                return SourceSnapshot(
                    key=snapshot_key,
                    name=snapshot_name,
                    state="partial",
                    summary="Glassdoor rate limit reached on RapidAPI.",
                    warning="RapidAPI quota is temporarily exhausted for the Glassdoor connector.",
                    updated_at=_now_iso(),
                )
            if status_code == 403:
                return SourceSnapshot(
                    key=snapshot_key,
                    name=snapshot_name,
                    state="offline",
                    summary="Glassdoor access is blocked by the provider.",
                    warning="RapidAPI denied access to the configured Glassdoor endpoint.",
                    updated_at=_now_iso(),
                )
            raise
        payload = response.json()
        items = _extract_records(payload)[: self.settings.glassdoor_limit]
        records = [
            _pick_fields(
                {
                    "company": item.get("company_name") or item.get("name") or item.get("employer_name"),
                    "rating": item.get("rating") or item.get("overall_rating"),
                    "reviews": item.get("review_count") or item.get("reviews_count"),
                    "industry": item.get("industry"),
                    "location": item.get("location") or item.get("headquarters"),
                },
            )
            for item in items
        ]

        return SourceSnapshot(
            key=snapshot_key,
            name=snapshot_name,
            state="online" if records else "partial",
            summary=f"Fetched {len(records)} live Glassdoor company results for '{active_query}'.",
            count=len(records),
            records=records,
            updated_at=_now_iso(),
        )

    async def fetch_stackexchange(self, client: httpx.AsyncClient) -> SourceSnapshot:
        params = {
            "order": "desc",
            "sort": "activity",
            "site": self.settings.stackexchange_site,
            "pagesize": self.settings.stackexchange_pagesize,
            "q": self.settings.stackexchange_query,
        }
        if self.settings.stackexchange_tag_list:
            params["tagged"] = ";".join(self.settings.stackexchange_tag_list)
        if self.settings.stackexchange_key:
            params["key"] = self.settings.stackexchange_key

        response = await client.get("https://api.stackexchange.com/2.3/search/advanced", params=params)
        response.raise_for_status()
        items = response.json().get("items", [])
        if not items and params.get("tagged"):
            params.pop("tagged", None)
            retry_response = await client.get(
                "https://api.stackexchange.com/2.3/search/advanced",
                params=params,
            )
            retry_response.raise_for_status()
            items = retry_response.json().get("items", [])
        records = [
            {
                "title": item.get("title"),
                "link": item.get("link"),
                "score": item.get("score"),
                "tags": item.get("tags"),
            }
            for item in items[: self.settings.stackexchange_pagesize]
        ]

        return SourceSnapshot(
            key="stackexchange",
            name="Stack Exchange",
            state="online" if records else "partial",
            summary=f"Fetched {len(records)} live engineering/forum results from Stack Exchange.",
            count=len(records),
            records=records,
            updated_at=_now_iso(),
        )

    async def fetch_reddit(self, client: httpx.AsyncClient) -> SourceSnapshot:
        if not (
            self.settings.reddit_client_id
            and self.settings.reddit_client_secret
            and self.settings.reddit_username
            and self.settings.reddit_password
        ):
            return SourceSnapshot(
                key="reddit",
                name="Reddit",
                state="offline",
                summary="Reddit feed is waiting for OAuth credentials.",
                warning="Reddit blocks the unauthenticated public JSON path used by this adapter.",
                updated_at=_now_iso(),
            )

        records: list[dict[str, Any]] = []
        headers = {"User-Agent": self.settings.reddit_user_agent}
        base_url = "https://www.reddit.com"

        basic_auth = base64.b64encode(
            f"{self.settings.reddit_client_id}:{self.settings.reddit_client_secret}".encode("utf-8"),
        ).decode("utf-8")
        token_response = await client.post(
            "https://www.reddit.com/api/v1/access_token",
            headers={
                "Authorization": f"Basic {basic_auth}",
                "User-Agent": self.settings.reddit_user_agent,
            },
            data={
                "grant_type": "password",
                "username": self.settings.reddit_username or "",
                "password": self.settings.reddit_password or "",
            },
        )
        try:
            token_response.raise_for_status()
        except httpx.HTTPStatusError as exc:
            status_code = exc.response.status_code
            if status_code in {401, 403}:
                return SourceSnapshot(
                    key="reddit",
                    name="Reddit",
                    state="offline",
                    summary="Reddit OAuth authentication failed.",
                    warning="Verify the Reddit app credentials and account permissions.",
                    updated_at=_now_iso(),
                )
            raise

        access_token = token_response.json().get("access_token")
        if access_token:
            headers = {
                "Authorization": f"Bearer {access_token}",
                "User-Agent": self.settings.reddit_user_agent,
            }
            base_url = "https://oauth.reddit.com"

        for subreddit in self.settings.subreddit_list[:3]:
            response = await client.get(
                f"{base_url}/r/{subreddit}/hot.json",
                headers=headers,
                params={"limit": max(self.settings.reddit_limit // max(len(self.settings.subreddit_list[:3]), 1), 2)},
            )
            try:
                response.raise_for_status()
            except httpx.HTTPStatusError as exc:
                status_code = exc.response.status_code
                if status_code in {401, 403, 429}:
                    return SourceSnapshot(
                        key="reddit",
                        name="Reddit",
                        state="partial" if records else "offline",
                        summary="Reddit restricted the current API request.",
                        warning="OAuth access was denied or rate limited by Reddit.",
                        count=len(records),
                        records=records[: self.settings.reddit_limit],
                        updated_at=_now_iso(),
                    )
                raise
            children = response.json().get("data", {}).get("children", [])
            for child in children[:2]:
                data = child.get("data", {})
                records.append(
                    {
                        "subreddit": subreddit,
                        "title": data.get("title"),
                        "score": data.get("score"),
                        "url": data.get("url"),
                    }
                )

        return SourceSnapshot(
            key="reddit",
            name="Reddit",
            state="online" if records else "partial",
            summary=f"Fetched {len(records)} live Reddit posts across {', '.join(self.settings.subreddit_list[:3])}.",
            count=len(records),
            records=records[: self.settings.reddit_limit],
            updated_at=_now_iso(),
        )

    async def fetch_catalogs(self, client: httpx.AsyncClient) -> SourceSnapshot:
        headers = {"User-Agent": self.settings.sec_user_agent}
        records: list[dict[str, Any]] = []
        online_count = 0
        warnings: list[str] = []

        for vendor, url in self.settings.catalog_targets.items():
            try:
                response = await client.get(url, headers=headers)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, "html.parser")
                candidates = []
                for selector in ("h1", "h2", "h3", "title"):
                    for node in soup.select(selector):
                        text = " ".join(node.get_text(" ", strip=True).split())
                        if len(text) >= 8 and text not in candidates:
                            candidates.append(text)
                        if len(candidates) >= 4:
                            break
                    if len(candidates) >= 4:
                        break
                if candidates:
                    online_count += 1
                else:
                    warnings.append(f"{vendor} returned no parseable headings.")
                records.append({"vendor": vendor, "url": url, "samples": candidates[:3]})
            except httpx.HTTPStatusError as exc:
                status_code = exc.response.status_code
                logger.warning(
                    "Catalog scrape failed for %s with HTTP %s: %s",
                    vendor,
                    status_code,
                    _describe_exception(exc),
                )
                if status_code in {401, 403}:
                    warnings.append(f"{vendor} blocked automated catalog access.")
                elif status_code == 429:
                    warnings.append(f"{vendor} rate limited the catalog scraper.")
                else:
                    warnings.append(f"{vendor} catalog returned HTTP {status_code}.")
                records.append({"vendor": vendor, "url": url, "samples": []})
            except httpx.TimeoutException as exc:
                logger.warning("Catalog scrape timed out for %s: %s", vendor, _describe_exception(exc))
                warnings.append(f"{vendor} catalog request timed out.")
                records.append({"vendor": vendor, "url": url, "samples": []})
            except Exception as exc:
                logger.warning(
                    "Catalog scrape failed for %s (%s): %s",
                    vendor,
                    exc.__class__.__name__,
                    _describe_exception(exc),
                )
                warnings.append(f"{vendor} catalog was temporarily unavailable.")
                records.append({"vendor": vendor, "url": url, "samples": []})

        state = "online" if online_count == len(self.settings.catalog_targets) else "partial"
        return SourceSnapshot(
            key="catalogs",
            name="Equipment Catalogs",
            state=state,
            summary=f"Scraped {online_count} of {len(self.settings.catalog_targets)} configured catalog sources.",
            count=len(records),
            records=records,
            warning="; ".join(warnings) if warnings else None,
            updated_at=_now_iso(),
        )

    def _assemble_dashboard(self, snapshots: list[SourceSnapshot]) -> DashboardPayload:
        stack_status = [
            StackStatusItem(name=_status_name(snapshot), state=snapshot.state)
            for snapshot in snapshots
        ]
        live_feed = [
            FeedEntry(
                source=_status_name(snapshot),
                tone=_tone_for_snapshot(snapshot),
                text=_display_text(snapshot),
            )
            for snapshot in snapshots
        ][:8]

        source_cards = [
            SourceCard(
                name=snapshot.name,
                volume=_source_volume(snapshot),
                detail=_display_text(snapshot),
            )
            for snapshot in snapshots
        ]

        return DashboardPayload(
            fetched_at=_now_iso(),
            stackStatus=stack_status,
            liveFeed=live_feed,
            proposalHighlights=PROPOSAL_HIGHLIGHTS,
            datasetProfile=DATASET_PROFILE,
            sourceCards=source_cards,
            sourceSnapshots=snapshots,
        )
